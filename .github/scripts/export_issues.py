#!/usr/bin/env python3
"""Export GitHub issues to an Excel workbook.

Environment variables:
  GH_TOKEN           GitHub token (in Actions: secrets.GITHUB_TOKEN)
  GITHUB_REPOSITORY  "owner/repo" (set automatically by GitHub Actions)
  OUTPUT_FILE        Output path (default: issues.xlsx)
"""

import os
import sys
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

API = "https://api.github.com"


def fetch_all_issues(repo: str, token: str | None) -> list[dict]:
    headers = {"Accept": "application/vnd.github+json",
               "X-GitHub-Api-Version": "2022-11-28"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    issues, page = [], 1
    while True:
        r = requests.get(
            f"{API}/repos/{repo}/issues",
            headers=headers,
            params={"state": "all", "per_page": 100, "page": page},
            timeout=30,
        )
        r.raise_for_status()
        batch = r.json()
        if not batch:
            break
        # The issues endpoint also returns pull requests; skip them.
        issues.extend(it for it in batch if "pull_request" not in it)
        page += 1
    return issues


def iso_to_local(s: str | None) -> str:
    if not s:
        return ""
    return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")


def build_workbook(repo: str, issues: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"

    headers = ["#", "Title", "State", "Labels", "Assignees",
               "Milestone", "Author", "Created", "Updated", "Closed", "URL"]
    widths = [8, 60, 10, 28, 20, 16, 16, 17, 17, 17, 45]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")
    body_font = Font(name="Arial")

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    for row, it in enumerate(sorted(issues, key=lambda x: x["number"], reverse=True), start=2):
        values = [
            it["number"],
            it["title"],
            it["state"],
            ", ".join(l["name"] for l in it.get("labels", [])),
            ", ".join(a["login"] for a in it.get("assignees", [])),
            (it.get("milestone") or {}).get("title", ""),
            (it.get("user") or {}).get("login", ""),
            iso_to_local(it.get("created_at")),
            iso_to_local(it.get("updated_at")),
            iso_to_local(it.get("closed_at")),
            it["html_url"],
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = body_font
            if col == 2:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    last_row = max(len(issues) + 1, 2)
    ws.auto_filter.ref = f"A1:K{last_row}"
    ws.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 22
    rows = [
        ("Repository", repo),
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("Total issues", f"=COUNTA(Issues!A2:A{last_row})" if issues else 0),
        ("Open", f'=COUNTIF(Issues!C2:C{last_row},"open")' if issues else 0),
        ("Closed", f'=COUNTIF(Issues!C2:C{last_row},"closed")' if issues else 0),
    ]
    for r, (k, v) in enumerate(rows, start=1):
        summary.cell(row=r, column=1, value=k).font = Font(name="Arial", bold=True)
        summary.cell(row=r, column=2, value=v).font = Font(name="Arial")

    wb.save(path)


def main() -> None:
    repo = os.environ.get("GITHUB_REPOSITORY")
    if not repo:
        sys.exit("GITHUB_REPOSITORY is not set (expected owner/repo)")
    token = os.environ.get("GH_TOKEN")
    out = os.environ.get("OUTPUT_FILE", "issues.xlsx")

    issues = fetch_all_issues(repo, token)
    build_workbook(repo, issues, out)
    print(f"Exported {len(issues)} issues from {repo} -> {out}")


if __name__ == "__main__":
    main()
